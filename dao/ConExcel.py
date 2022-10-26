from openpyxl import Workbook, workbook
from openpyxl.drawing.image import Image

class MetodosExcel:

    def __init__(self) -> None:
        self._book = None
        self._hojaActiva = None

    @property
    def book(self):
        return self._book

    @book.setter
    def book(self, book):
        self._book = book

    @property
    def hojaDefault(self):
        return self.hojaDefault

    @hojaDefault.setter
    def hojaDefault(self, hojaDefault):
        self.hojaDefault = hojaDefault

# ------------------------------

    def crearHoja(self, nameSheet):
        sheet = self._book.create_sheet(nameSheet)
        return sheet

    def hojaDefault(self):
        sheetDefault = self.book.active
        self.hojaDefault = sheetDefault
        # return sheetDefault

    def crearBook(self):
        self.book = Workbook()

    def guardarArchivo(self, nameFile):
        self.book.save(nameFile)
        print("guardando")

    def crearHeadersExcel(self):
        # crear datos colegio profesor
        self.hojaDefault.merge_cells(
            start_row=1, start_column=4, end_row=3, end_column=17)
        self.hojaDefault['D1'] = "COLEGIO BILINGÜE LIDIA SALMAN DE VARGAS"

        self.hojaDefault.merge_cells('D4:E4')
        self.hojaDefault['D4'] = "Grado"

        self.hojaDefault.merge_cells('D5:E5')
        self.hojaDefault['D5'] = "Docente"

        self.hojaDefault.merge_cells('F4:M4')
        self.hojaDefault['F4'] = "BACHILLERATO"
        self.hojaDefault.merge_cells('F5:M5')
        self.hojaDefault['F5'] = "Wiliam Pedorro"

        self.hojaDefault.merge_cells('N4:O4')
        self.hojaDefault['N4'] = "Asignatura"
        self.hojaDefault.merge_cells('N5:O5')
        self.hojaDefault['N5'] = "Firma"

        self.hojaDefault.merge_cells('P4:Q4')
        self.hojaDefault['P4'] = "iNFORMATICA"

        # TABLA NOTAS
        self.hojaDefault['A8'] = "N°"
        self.hojaDefault['B8'] = "NIE"
        # COMBINAR C-E
        self.hojaDefault.merge_cells('C8:E8')
        self.hojaDefault['C8'] = "ALUMNO/A"

        # CALIFICACIONES
        # COMBINAR F-L
        self.hojaDefault.merge_cells('F7:L7')
        self.hojaDefault['F7'] = "CALIFICACIONES"
        # fin cabecerar
        self.hojaDefault['F8'] = "1 Act"
        self.hojaDefault['G8'] = "35%"
        self.hojaDefault['H8'] = "2 Act"
        self.hojaDefault['I8'] = "35%"
        self.hojaDefault['J8'] = "Examen"
        self.hojaDefault['K8'] = "30%"
        self.hojaDefault['L8'] = "Prom"
        # FIN--CALIFICACIONES
        # Competencia ciudadanas
        # self.hojaDefault['M7']= "N°"
        # self.hojaDefault['N7']= "N°"
        # self.hojaDefault['O7']= "N°"
        # self.hojaDefault['P7']= "N°"
        # self.hojaDefault['Q7']= "N°"
        # NOTAS DE ESTUDIANTES

    def crearPlantillaNotas(self):
        self.crearHeadersExcel()
        counN = 7
        for no in range(1, 30):
            counN = counN+1
            self.hojaDefault[f'A{counN}'] = no
    
    def getTrimestre(self,periodo):
        trim=''
        if periodo== 'I':
            trim="PRIMER TRIMESTRE"
        if periodo== 'II':
            trim="SEGUNDO TRIMESTRE"
        if periodo== 'III':
            trim="TERCERO TRIMESTRE"
        if periodo== 'IIII':
            trim="CUARTO TRIMESTRE"
        return trim


    def crearPlantillaNotasExcel(self, calificaciones,grado,profesor,asignatura,periodo):
        self.crearHeadersExcel()
        counN = 9
        # for no in range(1,30):
        #     counN= counN+1
        #     self.hojaDefault[f'A{counN}']= no
        #SECCION IMAGEN
        img= Image("./assets/img/logo.png")
        img.height=125.5
        img.width=125.5
        self.hojaDefault.merge_cells(f'B1:C1')
        self.hojaDefault.add_image(img,'B1')
        
        self.hojaDefault.merge_cells(f'P4:Q4')
        self.hojaDefault['P4'] = asignatura

        self.hojaDefault.merge_cells(f'F4:M4')
        self.hojaDefault['F4'] = grado

        self.hojaDefault.merge_cells(f'F5:M5')
        self.hojaDefault['F5'] = profesor
        #TEXTO X TRIMESTRE
        self.hojaDefault.merge_cells(f'D6:L6')
        
        self.hojaDefault['D6'] = self.getTrimestre(periodo)
        
        counter=1
        for (index_row, row) in enumerate(calificaciones):
            nota1=0
            nota2=0
            nota3=0
            promedio=0

            #Número index
            self.hojaDefault[f'A{counN}']= counter
            #NIE
            self.hojaDefault.merge_cells(f'C{counN}:E{counN}')
            # AGREGAR NOMBRE
            self.hojaDefault[f'C{counN}']= str(row[2])
            #AGREGAR NOTA 1
            self.hojaDefault[f'F{counN}']= float(row[3])
            self.hojaDefault[f'G{counN}'.format(row)].number_format= '##.0' ##FORMATO 1 DECIMAL NOTA 1
            nota1=(float(row[3])*0.35)
            self.hojaDefault[f'G{counN}']= nota1

            #AGREGAR NOTA 2
            self.hojaDefault[f'H{counN}']= float(row[5])
            self.hojaDefault[f'I{counN}'.format(row)].number_format= '##.0' ##FORMATO 1 DECIMAL NOTA 2
            nota2=(float(row[5])*0.35)
            self.hojaDefault[f'I{counN}']= nota2

            #AGREGAR NOTA 3
            self.hojaDefault[f'J{counN}']= float(row[7])
            self.hojaDefault[f'K{counN}'.format(row)].number_format= '##.0' ##FORMATO 1 DECIMAL NOTA 3
            nota3=(float(row[7])*0.30)
            self.hojaDefault[f'K{counN}']= nota3 

            #PROMEDIO
            promedio=nota1+nota2+nota3
            self.hojaDefault[f'L{counN}'.format(row)].number_format= '##.0' ##FORMATO 1 DECIMAL PROMEDIO
            self.hojaDefault[f'L{counN}']= f"={promedio}"
            counN+=1
            counter+=1
